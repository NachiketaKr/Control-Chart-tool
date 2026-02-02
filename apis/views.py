import encodings
from tkinter import font
from numpy import NaN, fromstring
from .models import DataValue, Meter
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import HttpResponseRedirect, JsonResponse
from django.core import serializers
from .helper import *
import pandas as pd
from .serializers import DataValueSerializer, ExcelFilesSerializer, MeterSerializer
from django.http import HttpResponse
from django.shortcuts import render
from .models import ExcelFiles
from tablib import Dataset
import openpyxl
import csv
import datetime
import xlwt
@api_view(['GET'])
def getRoutes(request):
    routes=[
        {
            'Endpoint':'/health/score',
            'method':'GET',
            'body':{'value': ''},
            'description':'returns the health score for the given data values'
        },
        {
            'Endpoint':'/limits',
            'method':'GET',
            'body':None,
            'description':'returns the limits that is currently stored in the server '
        },
        {
            'Endpoint':'/data',
            'method':'GET',
            'body':None,
            'description':'returns all the data stored in the system'
        },
        {
            'Endpoint':'/data/create',
            'method':'POST',
            'body':{'body': ''},
            'description':'creates a new data entered by the user'
        },
        {
            'Endpoint':'/train/controlLimitType',
            'method':'PUT',
            'body':{
                'lcl':'',
                'ucl':''
            },
            'description':'the body is required only when we need to input specification limits'
        },


    ]
    return Response(routes)

@api_view(['POST'])
def getHealthScore(request):
    meterSet=Meter.objects.all()[0:30]
    dataMonth=[]
    for values in meterSet.iterator():
        dataMonth.append(values.value)
    dataValues=DataValue.objects.get(id=1)
    dataPoint=float(request.data['value'])
    limits=dataValues.limits
    coef=3
    lcl=round(float(limits[0]),3)
    cl=round(float(limits[(len(limits)-1)//2]),3)
    ucl=round(float(limits[len(limits)-1]),3)
    flag=[0,0,0,0,0]
    size_limits = len(limits)
    #---------------------Test for Rule 1-------------------_#
    if(dataPoint > limits[size_limits-1] or dataPoint < limits[0]):
        flag[0] = 1
    #---------------------Test for Rule 2-------------------_#
    nearUCL = test_near(dataMonth,limits,dataPoint,"UCL")
    nearLCL = test_near(dataMonth,limits,dataPoint,"LCL")
    if(nearLCL == 1 or nearUCL==1):
        flag[1]=1
    #---------------------Test for Rule 3-------------------_#
    checkUC = test_bUC_bLC(dataMonth,limits,dataPoint,"bUC")
    checkLC = test_bUC_bLC(dataMonth,limits,dataPoint,"bLC")
    if(checkUC == 1 or checkLC==1):
        flag[2]=1
    #---------------------Test for Rule 4-------------------_#
    checkTrend = test_trend_ud(dataMonth,limits,dataPoint)
    if(checkTrend==1 ):
        flag[3]=1
    
    healthscore=dataValues.healthScore
    score=100
    for i in range(len(healthscore)):
        if(flag[i]==1):
            score=healthscore[i]
            break

    scoreJson={
            'healthScore': score,  
        }
    return Response(scoreJson)

@api_view(['PUT'])
def training(request, choice):
    coef=3
    meterSet=Meter.objects.all()[0:30]
    dataMonth=[]
    for values in meterSet.iterator():
        dataMonth.append(values.value)
    limits=[]
    controlLimitType=1
    if(choice=='1'):
        lcl=float(request.data['lcl'])
        ucl=float(request.data['ucl'])
        limits=specification_limits_generator(lcl,ucl, coef)
        controlLimitType=1
    elif(choice=='2'):
        limits=limits_generator(dataMonth, coef)
        controlLimitType=2
    dataValues=DataValue.objects.get(id=1)
    values={
        'id':dataValues.id,
        'limits':limits,
        'healthScore':dataValues.healthScore,
        'controlLimitType':controlLimitType
    }
    dataValuesSerializer=DataValueSerializer(dataValues, data=values)
    if(dataValuesSerializer.is_valid()):
        dataValuesSerializer.save()
    lcl=round(float(limits[0]),3)
    cl=round(float(limits[(len(limits)-1)//2]),3)
    ucl=round(float(limits[len(limits)-1]),3)
    check_linearity = linearity(dataMonth)
    check_normality = normality(dataMonth)
    return Response({
        'lcl':lcl,
        'cl':cl,
        'ucl':ucl,
        'limits':limits,
        'control_limit_type':controlLimitType,
        'linearity_check':check_linearity,
        'normality_check':check_normality
    })


@api_view(['GET'])
def getLimits(request):
    limits=[]
    dataValues=DataValue.objects.get(id=1)
    limits=dataValues.limits
    lcl=round(float(limits[0]),3)
    cl=round(float(limits[(len(limits)-1)//2]),3)
    ucl=round(float(limits[len(limits)-1]),3)
    controlLimitType=dataValues.controlLimitType
    return Response({
        'lcl':lcl,
        'ucl':ucl,
        'cl':cl,
        'control_limit_type':controlLimitType
    })

@api_view(['GET'])
def getData(request):
    meterSet=Meter.objects.all()[0:30]
    dataMonth=[]
    for values in meterSet.iterator():
        dataMonth.append(values.value)
    return Response({
        'Meter_Values':dataMonth
    })
@api_view(['POST'])
def postData(request):
    dataPoint=float(request.data['value'])
    value={
        'value':dataPoint
    }
    dataPointSerializer=MeterSerializer(data=value)
    if(dataPointSerializer.is_valid()):
        dataPointSerializer.save()
    return Response({
        'value':dataPoint
    })

@api_view(['PUT'])
def upload_file(request):
    file = request.FILES['file']
    file_name=request.data['file_name']
    print(file_name)
    # fileData={
    #     'file':file,
    #     'file_name':file_name
    # }
    # fileDataSerializer=ExcelFileSerializer(data=fileData)
    # if fileDataSerializer.is_valid():
    #     fileDataSerializer.save()
    coef=3
    wb=openpyxl.load_workbook(file)
    worksheet=wb['Sheet1']
    print(coef)
    dataMonth=[]
    for row in worksheet.iter_rows():
        for cell in row:
            dataMonth.append(float(cell.value))
    limits=[]
    controlLimitType=3
    limits=limits_generator(dataMonth, coef)
    controlLimitType=3
    dataValues=DataValue.objects.get(id=1)
    values={
        'id':dataValues.id,
        'limits':limits,
        'healthScore':dataValues.healthScore,
        'controlLimitType':controlLimitType
    }
    dataValuesSerializer=DataValueSerializer(dataValues, data=values)
    if(dataValuesSerializer.is_valid()):
        dataValuesSerializer.save()
    lcl=round(float(limits[0]),3)
    cl=round(float(limits[(len(limits)-1)//2]),3)
    ucl=round(float(limits[len(limits)-1]),3)
    check_linearity = linearity(dataMonth)
    check_normality = normality(dataMonth)
    return Response({
        'lcl':lcl,
        'cl':cl,
        'ucl':ucl,
        'limits':limits,
        'control_limit_type':controlLimitType,
        'linearity_check':check_linearity,
        'normality_check':check_normality
    })

@api_view(['GET'])
def exportCSV(request):
    response=HttpResponse(content_type='text/csv')
    response['Content_Disposition']='attachment;filename=Process Information'+str(datetime.datetime.now())+'.csv'

    writer=csv.writer(response)
    writer.writerow(['Date', 'Value'])
    datavalues=Meter.objects.filter(owner=request.user)

    for datavalue in datavalues:
        writer.writerow([
            datavalue.created, datavalue.value
        ])
    return response

@api_view(['GET'])
def exportExcel(request):
    response=HttpResponse(content_type='application/ms-excel')
    response['Content_Disposition']='attachment;filename=Process Information'+str(datetime.datetime.now())+'.xls'
    wb=xlwt.Workbook(encoding='utf-8')
    ws=wb.add_sheet('DataValues')
    row_num=0
    font_style=xlwt.XFStyle()
    font_style.font.bold=True

    columns=['Date','Values']
    for col_num in range(len(columns)):
        ws.write(row_num, col_num,columns[col_num], font_style)

    font_style=xlwt.XFStyle()
    rows=Meter.objects.filter(owner=request.user).values_list('created','value')
    for row in rows:
        row_num+=1

        for col_num in range(len(row)):
            ws.write(row_num, col_num,columns[col_num], font_style)
    wb.save(response)
    return response

    